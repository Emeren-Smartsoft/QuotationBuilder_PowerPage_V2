import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
PATH = r'C:\Users\SMARTSOFT\Downloads\Lenovo Server Pricelist Q4 2024.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=True)
def s(v):
    if v is None: return ''
    return str(v).strip().replace('\xa0', ' ')
def num(v):
    if v is None: return 0
    try: return round(float(v), 2)
    except: return 0
bases = []
def parse(sheet, header_off, mode):
    ws = wb[sheet]; rows = list(ws.iter_rows(values_only=True))
    lb=ls=lf=lfa=''
    for r in rows[header_off:]:
        if not r[4]: continue
        b=s(r[0]) or lb; sk=s(r[1]) or ls; ff=s(r[2]) or lf; fa=s(r[3]) or lfa
        lb,ls,lf,lfa=b,sk,ff,fa
        if mode=='v1' or mode=='v2':
            bases.append({'brand':b or 'INTEL','socket':sk,'formFactor':ff,'family':fa,'partNo':s(r[4]),
                'processor':s(r[5]),'ghz':s(r[6]),'cache':s(r[7]),'memory':s(r[8]),'hdd':s(r[9]),
                'backplane':s(r[10]),'raid':s(r[11]),'mgmt':s(r[12]),'others':s(r[13]),'warranty':s(r[14]),
                'eeup':num(r[15]),'generation':'V1' if mode=='v1' else 'V2'})
        else:
            bases.append({'brand':b or 'AMD','socket':sk,'formFactor':ff,'family':fa,'partNo':s(r[4]),
                'processor':s(r[5]),'ghz':s(r[6]),'cache':'','memory':s(r[7]),'hdd':s(r[8]),
                'backplane':s(r[9]),'raid':s(r[10]),'mgmt':s(r[11]),'others':s(r[12]),'warranty':s(r[13]),
                'eeup':num(r[14]),'generation':'AMD'})
parse('Server Intel',3,'v1'); parse('Server V2 Intel',3,'v2'); parse('Server AMD',1,'amd')
options = {}
for sn in ['Options ( Active)','Software & Service']:
    ws=wb[sn]; lt=''
    for r in list(ws.iter_rows(values_only=True))[1:]:
        pn=s(r[2]); 
        if not pn: continue
        t=s(r[0]) or lt; lt=t
        if pn in options: continue
        options[pn]={'partNo':pn,'type':t or 'Software','description':s(r[3]),'eeup':num(r[4])}
compat={}
def parse_mat(sn, opt_col, base_col_start):
    ws=wb[sn]; rows=list(ws.iter_rows(values_only=True))
    base_cols={}
    for ci in range(base_col_start, ws.max_column):
        v=s(rows[2][ci])
        if v: base_cols[ci]=v; compat.setdefault(v,set())
    for r in rows[3:]:
        opn=s(r[opt_col])
        if not opn: continue
        for ci,bpn in base_cols.items():
            cell=r[ci] if ci<len(r) else None
            if cell and s(cell).lower()=='x':
                compat.setdefault(bpn,set()).add(opn)
parse_mat('Options Compactibility Matrix',2,4)
parse_mat('V2 Options Compactability Matri',1,3)
data={'bases':bases,'options':options,'compat':{k:sorted(v) for k,v in compat.items() if v}}
with open(r'd:\VS_CopilotStudio_Agents\Quote Site\_tmp\serverData.json','w',encoding='utf-8') as f:
    json.dump(data,f,separators=(',',':'))
print('OK', len(bases), len(options), len(data['compat']))
